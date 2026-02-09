import csv
import io
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from uuid import UUID
from typing import Optional

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import UserRole
from app.crud import processing_run as run_crud
from app.crud import document_type as document_type_crud
from app.crud import user as user_crud
from app.schemas.processing_run import (
    ProcessingRunResponse,
    ProcessingRunDetailResponse,
    ProcessingRunListResponse,
    ProcessingStatus,
    ProcessingSource,
    ProcessedDocumentResponse,
    ExtractedField,
    ExtractedFieldUpdate,
    ProcessingRunDocumentTypeUpdate,
)

router = APIRouter(prefix="/processing-runs", tags=["processing-runs"])


@router.get("", response_model=ProcessingRunListResponse)
def list_processing_runs(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    user_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    effective_user_id = user_id if current_user.role == UserRole.ADMIN else current_user.id
    runs, total = run_crud.get_processing_runs(db, skip, limit, effective_user_id)
    return ProcessingRunListResponse(items=runs, total=total)


@router.get("/by-document-type/{document_type_id}", response_model=ProcessingRunListResponse)
def get_processing_runs_by_document_type(
    document_type_id: UUID,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    user_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    document_type = document_type_crud.get_document_type(db, document_type_id)
    if not document_type:
        raise HTTPException(status_code=404, detail="Document type not found")

    effective_user_id = user_id if current_user.role == UserRole.ADMIN else current_user.id
    runs, total = run_crud.get_processing_runs_by_document_type(db, document_type_id, skip, limit, effective_user_id)
    return ProcessingRunListResponse(items=runs, total=total)


@router.get("/{run_id}", response_model=ProcessingRunDetailResponse)
def get_processing_run(run_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    run = run_crud.get_processing_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    if current_user.role != UserRole.ADMIN and run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return run


@router.post("/{document_type_id}", response_model=ProcessingRunResponse, status_code=201)
def create_processing_run(
    document_type_id: UUID,
    source: ProcessingSource = ProcessingSource.MANUAL,
    trigger_name: str = None,
    user_id: Optional[UUID] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    document_type = document_type_crud.get_document_type(db, document_type_id)
    if not document_type:
        raise HTTPException(status_code=404, detail="Document type not found")

    resolved_user_id = current_user.id
    if current_user.role == UserRole.ADMIN and user_id:
        if not user_crud.get_user(db, user_id):
            raise HTTPException(status_code=404, detail="User not found")
        resolved_user_id = user_id

    return run_crud.create_processing_run(
        db,
        document_type_id,
        source,
        trigger_name,
        user_id=resolved_user_id,
    )


@router.delete("/{run_id}", status_code=204)
def delete_processing_run(run_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    run = run_crud.get_processing_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    if current_user.role != UserRole.ADMIN and run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    success = run_crud.delete_processing_run(db, run_id)
    if not success:
        raise HTTPException(status_code=404, detail="Processing run not found")
    return None


@router.patch("/{run_id}/status", response_model=ProcessingRunResponse)
def update_processing_run_status(
    run_id: UUID,
    status: ProcessingStatus,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    run = run_crud.get_processing_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    if current_user.role != UserRole.ADMIN and run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    run = run_crud.update_processing_run_status(db, run_id, status)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    return run


@router.patch("/{run_id}/document-type", response_model=ProcessingRunResponse)
def update_processing_run_document_type(
    run_id: UUID,
    payload: ProcessingRunDocumentTypeUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    run = run_crud.get_processing_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    if current_user.role != UserRole.ADMIN and run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    document_type = document_type_crud.get_document_type(db, payload.document_type_id)
    if not document_type:
        raise HTTPException(status_code=404, detail="Document type not found")

    run = run_crud.update_processing_run_document_type(db, run_id, payload.document_type_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    return run


@router.post("/{run_id}/mark-reviewed", response_model=ProcessingRunResponse)
def mark_run_as_reviewed(run_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    run = run_crud.get_processing_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    if current_user.role != UserRole.ADMIN and run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    run = run_crud.mark_run_as_reviewed(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    return run


@router.post("/{run_id}/cancel-review", response_model=ProcessingRunResponse)
def cancel_run_review(run_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    run = run_crud.get_processing_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    if current_user.role != UserRole.ADMIN and run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    run = run_crud.cancel_run_review(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    return run


@router.get("/documents/{document_id}", response_model=ProcessedDocumentResponse)
def get_processed_document(document_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    doc = run_crud.get_processed_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.ADMIN and doc.processing_run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return doc


@router.patch("/documents/{document_id}/status")
def update_document_status(
    document_id: UUID,
    status: ProcessingStatus,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    from app.models.processed_document import DocumentStatus
    doc_status = DocumentStatus(status.value)
    doc = run_crud.get_processed_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.ADMIN and doc.processing_run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    doc = run_crud.update_document_status(db, document_id, doc_status)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"status": "updated"}


@router.post("/documents/{document_id}/mark-reviewed")
def mark_document_reviewed(document_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    doc = run_crud.get_processed_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.ADMIN and doc.processing_run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    doc = run_crud.mark_document_reviewed(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return {
        **ProcessedDocumentResponse.from_orm(doc).dict(),
        "run_status": doc.processing_run.status.value if doc.processing_run else None,
    }


@router.post("/documents/{document_id}/cancel-review")
def cancel_document_review(document_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    doc = run_crud.get_processed_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.ADMIN and doc.processing_run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    doc = run_crud.cancel_document_review(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return {
        **ProcessedDocumentResponse.from_orm(doc).dict(),
        "run_status": doc.processing_run.status.value if doc.processing_run else None,
    }


@router.get("/documents/{document_id}/fields", response_model=list[ExtractedField])
def get_document_fields(document_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    doc = run_crud.get_processed_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.ADMIN and doc.processing_run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    return run_crud.get_extracted_fields_by_document(db, document_id)


@router.patch("/documents/{document_id}/fields", response_model=ProcessedDocumentResponse)
def update_extracted_field(
    document_id: UUID,
    field_data: ExtractedFieldUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    doc = run_crud.get_processed_document(db, document_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if current_user.role != UserRole.ADMIN and doc.processing_run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    doc = run_crud.update_extracted_field(db, document_id, field_data.field_index, field_data.value)
    if not doc:
        raise HTTPException(status_code=404, detail="Document or field not found")
    return doc


def _build_export_rows(run, export_keys: dict) -> tuple[list[str], list[list[str]]]:
    """Build header + rows from a processing run's documents."""
    # Collect all field names across documents
    all_field_names: list[str] = []
    seen: set[str] = set()
    for doc in run.documents:
        for field in (doc.extracted_fields or []):
            name = field.get("name", "")
            if name and name not in seen:
                all_field_names.append(name)
                seen.add(name)

    # Apply export_keys mapping: field_name -> export_key
    headers = ["Файл"]
    for name in all_field_names:
        headers.append(export_keys.get(name, name))

    rows: list[list[str]] = []
    for doc in run.documents:
        field_map = {f["name"]: f["value"] for f in (doc.extracted_fields or [])}
        row = [doc.filename]
        for name in all_field_names:
            row.append(str(field_map.get(name, "")))
        rows.append(row)

    return headers, rows


@router.get("/{run_id}/export")
def export_processing_run(
    run_id: UUID,
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    run = run_crud.get_processing_run(db, run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Processing run not found")
    if current_user.role != UserRole.ADMIN and run.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")

    document_type = run.document_type
    export_keys = (document_type.export_keys or {}) if document_type else {}
    type_name = document_type.name if document_type else "export"
    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in type_name)

    headers, rows = _build_export_rows(run, export_keys)

    if not rows:
        raise HTTPException(status_code=404, detail="No documents to export")

    if format == "csv":
        output = io.StringIO()
        output.write("\ufeff")  # BOM for Excel compatibility
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        content = output.getvalue().encode("utf-8")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
        )

    # Excel (xlsx)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = safe_name[:31]  # Excel sheet name max 31 chars

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="107572", end_color="107572", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    data_alignment = Alignment(vertical="center", wrap_text=True)
    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_data in rows:
            val_len = len(str(row_data[col_idx - 1])) if col_idx - 1 < len(row_data) else 0
            max_len = max(max_len, val_len)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )
